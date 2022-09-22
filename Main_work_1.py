import openpyxl
import pandas as pd
from datetime import datetime, time, date


class data_person(object):
        @staticmethod
        def write():
            r = sheet.max_row
            cell = sheet.cell(row=r + 1, column=1)
            cell.value = input("Введите имя: ")
            cell = sheet.cell(row=r + 1, column=2)
            cell.value = input("Введите фамилию: ")
            cell = sheet.cell(row=r + 1, column=3)
            cell.value = input("Введите отчество: ")
            cell = sheet.cell(row=r + 1, column=4)
            cell.value = input("Введите дату рождения(ГГ-ММ-ДД): ")
            cell = sheet.cell(row=r + 1, column=5)
            cell.value = input("Введите дату смерти(ГГ-ММ-ДД): ")
            cell = sheet.cell(row=r + 1, column=6)
            cell.value = input("Введите пол(М/Ж): ")
            wb.save("Test.xlsx")

        @staticmethod
        def load_data():
            try:
                wb = openpyxl.Workbook()
                wb.create_sheet(title="Sheet1", index=0)
                wb.save("Test.xlsx")
                file_name = input("Введите назваие файла: ")
                data_for_load = pd.read_excel(file_name)
                writer = pd.ExcelWriter("Test.xlsx",  engine='openpyxl')
                data_for_load.to_excel(writer, index=False)
                writer.save()
                print("Данные внесены ")
            except FileNotFoundError:
                print(f"Файл `{file_name}` не найден")

        @staticmethod
        def find_data():
            text = input("   Введите имя для поиска: ")
            persons = pd.read_excel("Test.xlsx", sheet_name="Sheet1")
            filter_person = persons["name"].str.contains(text)
            filter_person_1 = persons["surname"].str.contains(text)
            filter_person_2 = persons["second_name"].str.contains(text)
            print(persons.loc[filter_person | filter_person_1 | filter_person_2])


        @staticmethod
        def person_age():
            text = input("   Введите имя для поиска: ")
            today = date.today()
            persons = pd.read_excel("Test.xlsx", sheet_name="Sheet1")
            filter_person = persons["name"].str.contains(text)
            filter_person_1 = persons["surname"].str.contains(text)
            filter_person_2 = persons["second_name"].str.contains(text)
            person_to_age = persons.loc[filter_person | filter_person_1 | filter_person_2]
            name_person = person_to_age[["name", "surname", "second_name"]]

            def age_die(age_die_person):
                if dieday_p_dt.month < birthday_p_dt.month:
                    age_die_person -= 1
                elif dieday_p_dt.month == birthday_p_dt.month and dieday_p_dt.day < birthday_p_dt.day:
                    age_die_person -= 1
                return f"{name_person} прожил {age_die_person} лет/года"

            def age(age_person):
                if today.month < birthday_p_dt.month:
                    age_person -= 1
                elif today.month == birthday_p_dt.month and today.day < birthday_p_dt.day:
                    age_person -= 1
                return f"{name_person} сейчас {age_person} лет/года"
            try:
                birthday_person = person_to_age["birthday"]
                birthday_person_str = birthday_person.to_string()
                birthday_p_dt = datetime.strptime(birthday_person_str[-10:], "%Y-%m-%d")
                dieday_person = person_to_age["dieday"]
                dieday_person_str = dieday_person.to_string()
                dieday_p_dt = datetime.strptime(dieday_person_str[-10:], "%Y-%m-%d")
                age_die_person = dieday_p_dt.year - birthday_p_dt.year
                print(age_die(age_die_person))
            except ValueError:
                birthday_person = person_to_age["birthday"]
                birthday_person_str = birthday_person.to_string()
                birthday_p_dt = datetime.strptime(birthday_person_str[-10:], "%Y-%m-%d")
                age_person = datetime.now().year - birthday_p_dt.year
                print(age(age_person))


wb = openpyxl.Workbook()
wb.create_sheet(title="Sheet1", index=0)
sheet = wb["Sheet1"]
name_colums = ["name", "surname", "second_name", "birthday", "dieday", "sex"]

for i in range(0, 6):
    cell = sheet.cell(row=1, column=i+1)
    cell.value = name_colums[i]

wb.save("Test.xlsx")
