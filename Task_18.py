# Прочитать сохранённый csv-файл из задания №17 и сохранить данные в excel-файл,
# кроме возраста – столбец с этими данными не нужен.
# К заданию прикреплён пример как должно выглядеть содержания итогового файла.
import openpyxl
import csv


new_data = []

with open("friends_1.csv" ) as f:
    file_reader = csv.reader(f)
    for row in file_reader:
        new_data.append(row)


name_row, first, second, third, forth, fiftt, six = new_data
pers_nums = ['', 'Person 1', 'Person 2', 'Person 3', 'Person 4', 'Person 5', 'Person 6']
wb = openpyxl.Workbook()
sheet = wb["Sheet"]

for i in range(1, 7):
    cell = sheet.cell(row=1, column=i+1)
    cell.value = pers_nums[i]

for row_index, row in enumerate((name_row, first, second, third, forth, fiftt, six)):
    for col_index, value in enumerate(row):
        cell = sheet.cell(column=row_index+1, row=2+col_index)
        cell.value = value

sheet.delete_rows(4)
wb.save("task_18.xlsx")

